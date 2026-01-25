# -*- coding: utf-8 -*-
"""
assessment_importer.py - 评估数据导入器

功能:
1. 从原始目录批量导入 HRV 生理数据和心理测评数据
2. 解析 Excel 文件提取结构化数据
3. 解析 PDF 报告提取关键指标
4. 生成标准化 JSON 格式
5. 自动触发处方引擎生成建议

支持的数据源:
- 海棠心智设备导出的 Excel 文件 (生理/心理测评数据)
- 心理健康测评 PDF 报告

使用方法:
    python scripts/assessment_importer.py --source "D:\情绪评估报告\情绪评估报告"
    python scripts/assessment_importer.py --source "D:\情绪评估报告\情绪评估报告" --generate-prescriptions
"""

import json
import re
import shutil
import argparse
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from dataclasses import dataclass, asdict
import uuid

# Excel 解析
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# PDF 解析
try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

# 备选 PDF 解析
try:
    from PyPDF2 import PdfReader
    PYPDF2_AVAILABLE = True
except ImportError:
    PYPDF2_AVAILABLE = False


# ============ 数据结构 ============

@dataclass
class HeartRateData:
    """心率数据"""
    avg: float = 0.0
    min: float = 0.0
    max: float = 0.0
    unit: str = "bpm"


@dataclass
class HRVData:
    """心率变异性数据"""
    sdnn: float = 0.0       # 心脏调节能力
    rmssd: float = 0.0      # 恢复能力
    lf: float = 0.0         # 低频功率
    hf: float = 0.0         # 高频功率
    lf_hf_ratio: float = 0.0  # LF/HF 比值
    unit: str = "ms"


@dataclass
class PhysioData:
    """生理数据"""
    heart_rate: HeartRateData = None
    hrv: HRVData = None
    autonomic_balance: str = "unknown"  # balanced, sympathetic, parasympathetic

    def __post_init__(self):
        if self.heart_rate is None:
            self.heart_rate = HeartRateData()
        if self.hrv is None:
            self.hrv = HRVData()


@dataclass
class PsychData:
    """心理数据"""
    stress_index: float = 50.0      # 压力指数 (0-100)
    fatigue_index: float = 50.0     # 疲劳指数 (0-100)
    mood_index: float = 50.0        # 心情指数 (0-100)
    anxiety_score: float = 30.0     # 焦虑评分
    composite_score: float = 70.0   # 综合得分 (0-100)
    emotion_distribution: Dict[str, float] = None

    def __post_init__(self):
        if self.emotion_distribution is None:
            self.emotion_distribution = {}


@dataclass
class Assessment:
    """评估记录"""
    assessment_id: str = ""
    user_id: str = ""
    device_id: str = ""
    assessment_date: str = ""
    assessment_time: str = ""
    duration_minutes: int = 0
    device: str = "海棠心智-C2"

    physio_data: PhysioData = None
    psych_data: PsychData = None

    risk_level: str = "medium"
    risk_flags: List[str] = None
    source_files: Dict[str, str] = None

    def __post_init__(self):
        if self.physio_data is None:
            self.physio_data = PhysioData()
        if self.psych_data is None:
            self.psych_data = PsychData()
        if self.risk_flags is None:
            self.risk_flags = []
        if self.source_files is None:
            self.source_files = {}

    def to_dict(self) -> Dict[str, Any]:
        """转换为字典 (用于 JSON 序列化)"""
        return {
            "assessment_id": self.assessment_id,
            "user_id": self.user_id,
            "device_id": self.device_id,
            "assessment_date": self.assessment_date,
            "assessment_time": self.assessment_time,
            "duration_minutes": self.duration_minutes,
            "device": self.device,
            "physio_data": {
                "heart_rate": asdict(self.physio_data.heart_rate),
                "hrv": asdict(self.physio_data.hrv),
                "autonomic_balance": self.physio_data.autonomic_balance
            },
            "psych_data": {
                "stress_index": self.psych_data.stress_index,
                "fatigue_index": self.psych_data.fatigue_index,
                "mood_index": self.psych_data.mood_index,
                "anxiety_score": self.psych_data.anxiety_score,
                "composite_score": self.psych_data.composite_score,
                "emotion_distribution": self.psych_data.emotion_distribution
            },
            "risk_assessment": {
                "level": self.risk_level,
                "flags": self.risk_flags
            },
            "source_files": self.source_files
        }


# ============ 解析器 ============

class ExcelParser:
    """Excel 文件解析器"""

    @staticmethod
    def parse_physio_excel(file_path: Path) -> PhysioData:
        """解析生理测评数据 Excel"""
        if not OPENPYXL_AVAILABLE:
            print(f"[警告] openpyxl 未安装，无法解析 Excel: {file_path}")
            return PhysioData()

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            physio = PhysioData()

            # 遍历所有单元格查找关键指标
            for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=20):
                for cell in row:
                    if cell.value is None:
                        continue

                    cell_str = str(cell.value).strip()

                    # 匹配心率相关
                    if '平均心率' in cell_str or 'avg' in cell_str.lower():
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value:
                            physio.heart_rate.avg = float(next_cell.value)

                    elif '最小心率' in cell_str or 'min' in cell_str.lower():
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value:
                            physio.heart_rate.min = float(next_cell.value)

                    elif '最大心率' in cell_str or 'max' in cell_str.lower():
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value:
                            physio.heart_rate.max = float(next_cell.value)

                    # 匹配 HRV 相关
                    elif 'SDNN' in cell_str.upper() or '心脏调节' in cell_str:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value:
                            physio.hrv.sdnn = float(next_cell.value)

                    elif 'RMSSD' in cell_str.upper() or '恢复能力' in cell_str:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value:
                            physio.hrv.rmssd = float(next_cell.value)

                    elif 'LF/HF' in cell_str.upper() or '神经平衡' in cell_str:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value:
                            try:
                                physio.hrv.lf_hf_ratio = float(next_cell.value)
                            except:
                                pass

            # 判断自主神经平衡
            if physio.hrv.lf_hf_ratio > 0:
                if physio.hrv.lf_hf_ratio < 0.8:
                    physio.autonomic_balance = "parasympathetic"
                elif physio.hrv.lf_hf_ratio > 1.5:
                    physio.autonomic_balance = "sympathetic"
                else:
                    physio.autonomic_balance = "balanced"

            wb.close()
            return physio

        except Exception as e:
            print(f"[错误] 解析生理数据失败 {file_path}: {e}")
            return PhysioData()

    @staticmethod
    def parse_psych_excel(file_path: Path) -> PsychData:
        """解析心理测评数据 Excel"""
        if not OPENPYXL_AVAILABLE:
            print(f"[警告] openpyxl 未安装，无法解析 Excel: {file_path}")
            return PsychData()

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            psych = PsychData()

            # 遍历所有单元格查找关键指标
            for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=20):
                for cell in row:
                    if cell.value is None:
                        continue

                    cell_str = str(cell.value).strip()

                    if '压力' in cell_str and '指数' in cell_str:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value:
                            psych.stress_index = float(next_cell.value)

                    elif '疲劳' in cell_str and '指数' in cell_str:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value:
                            psych.fatigue_index = float(next_cell.value)

                    elif '心情' in cell_str and '指数' in cell_str:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value:
                            psych.mood_index = float(next_cell.value)

                    elif '综合' in cell_str and ('得分' in cell_str or '分数' in cell_str):
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value:
                            psych.composite_score = float(next_cell.value)

            wb.close()
            return psych

        except Exception as e:
            print(f"[错误] 解析心理数据失败 {file_path}: {e}")
            return PsychData()


class PDFParser:
    """PDF 报告解析器"""

    @staticmethod
    def parse_report_pdf(file_path: Path) -> Tuple[PhysioData, PsychData, Dict[str, Any]]:
        """
        解析心理健康测评报告 PDF

        Returns:
            (PhysioData, PsychData, meta_info)
        """
        physio = PhysioData()
        psych = PsychData()
        meta = {
            "assessment_date": "",
            "assessment_time": "",
            "duration_minutes": 0,
            "device": "海棠心智-C2"
        }

        text = PDFParser._extract_text(file_path)
        if not text:
            return physio, psych, meta

        # 解析元信息
        date_match = re.search(r'测评日期[：:]\s*(\d{4}-\d{2}-\d{2})', text)
        if date_match:
            meta["assessment_date"] = date_match.group(1)

        time_match = re.search(r'测评时间[：:]\s*(\d{1,2}:\d{2}\s*[-–]\s*\d{1,2}:\d{2})', text)
        if time_match:
            meta["assessment_time"] = time_match.group(1)

        duration_match = re.search(r'测评时长[：:]\s*(\d+)\s*分钟', text)
        if duration_match:
            meta["duration_minutes"] = int(duration_match.group(1))

        # 解析综合得分
        score_match = re.search(r'综合得分[：:]\s*(\d+)', text)
        if score_match:
            psych.composite_score = float(score_match.group(1))

        # 解析心理指标
        stress_match = re.search(r'综合压力[：:]*\s*(\d+)', text)
        if stress_match:
            psych.stress_index = float(stress_match.group(1))

        fatigue_match = re.search(r'综合疲劳[：:]*\s*(\d+)', text)
        if fatigue_match:
            psych.fatigue_index = float(fatigue_match.group(1))

        mood_match = re.search(r'综合心情[：:]*\s*(\d+)', text)
        if mood_match:
            psych.mood_index = float(mood_match.group(1))

        # 解析生理指标
        hr_avg_match = re.search(r'平均心率[：:]*\s*(\d+)\s*bpm', text)
        if hr_avg_match:
            physio.heart_rate.avg = float(hr_avg_match.group(1))

        hr_min_match = re.search(r'最小心率[：:]*\s*(\d+)\s*bpm', text)
        if hr_min_match:
            physio.heart_rate.min = float(hr_min_match.group(1))

        hr_max_match = re.search(r'最大心率[：:]*\s*(\d+)\s*bpm', text)
        if hr_max_match:
            physio.heart_rate.max = float(hr_max_match.group(1))

        sdnn_match = re.search(r'(?:心脏调节|SDNN)[：:]*\s*(\d+)\s*ms', text)
        if sdnn_match:
            physio.hrv.sdnn = float(sdnn_match.group(1))

        rmssd_match = re.search(r'(?:恢复能力|RMSSD)[：:]*\s*(\d+)\s*ms', text)
        if rmssd_match:
            physio.hrv.rmssd = float(rmssd_match.group(1))

        # 解析情绪分布
        calm_match = re.search(r'平静\s*(\d+)%', text)
        relaxed_match = re.search(r'松懈\s*(\d+)%', text)
        peaceful_match = re.search(r'安宁\s*(\d+)%', text)

        if calm_match:
            psych.emotion_distribution['calm'] = float(calm_match.group(1)) / 100
        if relaxed_match:
            psych.emotion_distribution['relaxed'] = float(relaxed_match.group(1)) / 100
        if peaceful_match:
            psych.emotion_distribution['peaceful'] = float(peaceful_match.group(1)) / 100

        # 判断神经平衡
        if '平衡' in text:
            physio.autonomic_balance = "balanced"
        elif '交感' in text:
            physio.autonomic_balance = "sympathetic"
        elif '副交感' in text:
            physio.autonomic_balance = "parasympathetic"

        return physio, psych, meta

    @staticmethod
    def _extract_text(file_path: Path) -> str:
        """从 PDF 提取文本"""
        text = ""

        # 优先使用 pdfplumber
        if PDFPLUMBER_AVAILABLE:
            try:
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"
                return text
            except Exception as e:
                print(f"[pdfplumber 错误] {e}")

        # 备选使用 PyPDF2
        if PYPDF2_AVAILABLE:
            try:
                reader = PdfReader(file_path)
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                return text
            except Exception as e:
                print(f"[PyPDF2 错误] {e}")

        print(f"[警告] 无可用的 PDF 解析库，请安装 pdfplumber 或 PyPDF2")
        return ""


# ============ 导入器 ============

class AssessmentImporter:
    """评估数据导入器"""

    def __init__(self, project_root: Path = None):
        """
        初始化导入器

        Args:
            project_root: 项目根目录
        """
        self.project_root = project_root or Path("D:/behavioral-health-project")
        self.data_dir = self.project_root / "data" / "assessments"
        self.raw_dir = self.data_dir / "raw"
        self.processed_dir = self.data_dir / "processed"

        # 确保目录存在
        self.raw_dir.mkdir(parents=True, exist_ok=True)
        self.processed_dir.mkdir(parents=True, exist_ok=True)
        (self.processed_dir / "users").mkdir(exist_ok=True)

    def import_from_directory(
        self,
        source_dir: Path,
        batch_name: str = None,
        copy_files: bool = True
    ) -> List[Assessment]:
        """
        从目录批量导入评估数据

        Args:
            source_dir: 源数据目录 (如 D:\情绪评估报告\情绪评估报告)
            batch_name: 批次名称 (默认使用当前日期)
            copy_files: 是否复制原始文件到项目目录

        Returns:
            导入的 Assessment 列表
        """
        source_dir = Path(source_dir)
        if not source_dir.exists():
            print(f"[错误] 源目录不存在: {source_dir}")
            return []

        batch_name = batch_name or datetime.now().strftime("%Y-%m-%d")
        batch_dir = self.raw_dir / f"batch_{batch_name}"

        print(f"\n{'='*60}")
        print(f"评估数据导入")
        print(f"{'='*60}")
        print(f"  源目录: {source_dir}")
        print(f"  批次: {batch_name}")
        print(f"  目标: {batch_dir}")

        assessments = []

        # 遍历用户目录
        user_dirs = [d for d in source_dir.iterdir() if d.is_dir()]
        print(f"  发现 {len(user_dirs)} 个用户目录")

        for user_dir in user_dirs:
            device_id_short = user_dir.name.upper()

            # 查找文件
            physio_file = None
            psych_file = None
            report_file = None

            for f in user_dir.glob("*"):
                fname = f.name
                if '生理测评' in fname and fname.endswith('.xlsx'):
                    physio_file = f
                elif '心理测评' in fname and fname.endswith('.xlsx'):
                    psych_file = f
                elif fname.endswith('.pdf'):
                    report_file = f

            if not any([physio_file, psych_file, report_file]):
                print(f"  [跳过] {device_id_short}: 无有效数据文件")
                continue

            # 解析数据
            assessment = self._parse_user_data(
                device_id_short,
                physio_file,
                psych_file,
                report_file
            )

            if assessment:
                # 复制文件到项目目录
                if copy_files:
                    self._copy_files_to_project(
                        assessment,
                        physio_file,
                        psych_file,
                        report_file,
                        batch_dir / device_id_short
                    )

                # 保存标准化 JSON
                self._save_assessment(assessment)

                assessments.append(assessment)
                print(f"  [导入] {assessment.user_id}: 综合得分={assessment.psych_data.composite_score}")

            # 递归处理子目录
            for sub_dir in user_dir.iterdir():
                if sub_dir.is_dir():
                    sub_assessments = self._process_subdirectory(sub_dir, batch_dir, copy_files)
                    assessments.extend(sub_assessments)

        print(f"\n导入完成: {len(assessments)} 条评估记录")

        # 更新索引
        self._update_index(assessments)

        return assessments

    def _process_subdirectory(self, sub_dir: Path, batch_dir: Path, copy_files: bool) -> List[Assessment]:
        """处理子目录"""
        assessments = []
        device_id_short = sub_dir.name.upper()

        physio_file = None
        psych_file = None
        report_file = None

        for f in sub_dir.glob("*"):
            fname = f.name
            if '生理测评' in fname and fname.endswith('.xlsx'):
                physio_file = f
            elif '心理测评' in fname and fname.endswith('.xlsx'):
                psych_file = f
            elif fname.endswith('.pdf'):
                report_file = f

        if any([physio_file, psych_file, report_file]):
            assessment = self._parse_user_data(
                device_id_short,
                physio_file,
                psych_file,
                report_file
            )

            if assessment:
                if copy_files:
                    self._copy_files_to_project(
                        assessment,
                        physio_file,
                        psych_file,
                        report_file,
                        batch_dir / device_id_short
                    )

                self._save_assessment(assessment)
                assessments.append(assessment)
                print(f"  [导入] {assessment.user_id}: 综合得分={assessment.psych_data.composite_score}")

        return assessments

    def _parse_user_data(
        self,
        device_id_short: str,
        physio_file: Path,
        psych_file: Path,
        report_file: Path
    ) -> Optional[Assessment]:
        """解析单个用户的数据"""
        assessment = Assessment()

        # 从文件名提取完整设备ID
        full_device_id = device_id_short
        for f in [physio_file, psych_file, report_file]:
            if f:
                match = re.search(r'[【\[]([A-F0-9]{12})[】\]]', f.name, re.IGNORECASE)
                if match:
                    full_device_id = match.group(1).upper()
                    break

        assessment.device_id = device_id_short
        assessment.user_id = full_device_id

        # 从文件名提取日期
        for f in [physio_file, psych_file, report_file]:
            if f:
                date_match = re.search(r'(\d{4})[.\-](\d{2})[.\-](\d{2})', f.name)
                if date_match:
                    assessment.assessment_date = f"{date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}"
                    break

        # 生成评估ID
        date_str = assessment.assessment_date.replace("-", "") if assessment.assessment_date else datetime.now().strftime("%Y%m%d")
        assessment.assessment_id = f"ASS-{date_str}-{device_id_short}"

        # 解析 PDF (优先，包含最完整的数据)
        if report_file and report_file.exists():
            physio_pdf, psych_pdf, meta = PDFParser.parse_report_pdf(report_file)

            # 合并 PDF 数据
            if meta.get("assessment_date"):
                assessment.assessment_date = meta["assessment_date"]
            if meta.get("assessment_time"):
                assessment.assessment_time = meta["assessment_time"]
            if meta.get("duration_minutes"):
                assessment.duration_minutes = meta["duration_minutes"]

            assessment.physio_data = physio_pdf
            assessment.psych_data = psych_pdf

        # 解析 Excel 文件 (补充/覆盖数据)
        if physio_file and physio_file.exists():
            physio_excel = ExcelParser.parse_physio_excel(physio_file)
            # 合并数据 (Excel 数据可能更精确)
            if physio_excel.heart_rate.avg > 0:
                assessment.physio_data.heart_rate = physio_excel.heart_rate
            if physio_excel.hrv.sdnn > 0:
                assessment.physio_data.hrv = physio_excel.hrv

        if psych_file and psych_file.exists():
            psych_excel = ExcelParser.parse_psych_excel(psych_file)
            # 合并数据
            if psych_excel.stress_index > 0:
                assessment.psych_data.stress_index = psych_excel.stress_index
            if psych_excel.fatigue_index > 0:
                assessment.psych_data.fatigue_index = psych_excel.fatigue_index
            if psych_excel.mood_index > 0:
                assessment.psych_data.mood_index = psych_excel.mood_index

        # 评估风险等级
        assessment.risk_level, assessment.risk_flags = self._assess_risk(assessment)

        return assessment

    def _assess_risk(self, assessment: Assessment) -> Tuple[str, List[str]]:
        """评估风险等级"""
        flags = []
        score = assessment.psych_data.composite_score

        # 基于综合得分
        if score < 60:
            level = "high"
            flags.append("综合得分低于60，急需关注")
        elif score < 80:
            level = "medium"
            flags.append("综合得分在61-80之间，需要注意")
        else:
            level = "low"

        # 基于 HRV
        if assessment.physio_data.hrv.sdnn < 30:
            level = "high"
            flags.append(f"SDNN={assessment.physio_data.hrv.sdnn}ms，心脏调节能力偏低")

        if assessment.physio_data.hrv.rmssd < 25:
            flags.append(f"RMSSD={assessment.physio_data.hrv.rmssd}ms，恢复能力偏低")

        # 基于心理指标
        if assessment.psych_data.stress_index > 70:
            flags.append(f"压力指数={assessment.psych_data.stress_index}，压力偏高")

        if assessment.psych_data.fatigue_index > 70:
            flags.append(f"疲劳指数={assessment.psych_data.fatigue_index}，疲劳明显")

        if assessment.psych_data.mood_index < 40:
            flags.append(f"心情指数={assessment.psych_data.mood_index}，情绪低落")

        return level, flags

    def _copy_files_to_project(
        self,
        assessment: Assessment,
        physio_file: Path,
        psych_file: Path,
        report_file: Path,
        target_dir: Path
    ):
        """复制原始文件到项目目录"""
        target_dir.mkdir(parents=True, exist_ok=True)

        if physio_file and physio_file.exists():
            shutil.copy2(physio_file, target_dir / "physio.xlsx")
            assessment.source_files["physio_xlsx"] = str(target_dir / "physio.xlsx")

        if psych_file and psych_file.exists():
            shutil.copy2(psych_file, target_dir / "psych.xlsx")
            assessment.source_files["psych_xlsx"] = str(target_dir / "psych.xlsx")

        if report_file and report_file.exists():
            shutil.copy2(report_file, target_dir / "report.pdf")
            assessment.source_files["report_pdf"] = str(target_dir / "report.pdf")

    def _save_assessment(self, assessment: Assessment):
        """保存标准化评估 JSON"""
        user_dir = self.processed_dir / "users" / assessment.user_id
        assessments_dir = user_dir / "assessments"
        assessments_dir.mkdir(parents=True, exist_ok=True)

        # 保存评估记录
        output_file = assessments_dir / f"{assessment.assessment_date}.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(assessment.to_dict(), f, ensure_ascii=False, indent=2)

        # 更新用户档案
        profile_file = user_dir / "profile.json"
        profile = {
            "user_id": assessment.user_id,
            "device_id": assessment.device_id,
            "first_assessment": assessment.assessment_date,
            "last_assessment": assessment.assessment_date,
            "total_assessments": 1
        }

        if profile_file.exists():
            with open(profile_file, 'r', encoding='utf-8') as f:
                existing = json.load(f)
                profile["first_assessment"] = existing.get("first_assessment", assessment.assessment_date)
                profile["total_assessments"] = existing.get("total_assessments", 0) + 1

        with open(profile_file, 'w', encoding='utf-8') as f:
            json.dump(profile, f, ensure_ascii=False, indent=2)

    def _update_index(self, assessments: List[Assessment]):
        """更新用户索引"""
        index_file = self.processed_dir / "index.json"
        index = {"users": {}, "last_updated": datetime.now().isoformat()}

        if index_file.exists():
            with open(index_file, 'r', encoding='utf-8') as f:
                index = json.load(f)

        for ass in assessments:
            if ass.user_id not in index["users"]:
                index["users"][ass.user_id] = {
                    "device_id": ass.device_id,
                    "assessments": []
                }

            if ass.assessment_date not in index["users"][ass.user_id]["assessments"]:
                index["users"][ass.user_id]["assessments"].append(ass.assessment_date)

        index["last_updated"] = datetime.now().isoformat()
        index["total_users"] = len(index["users"])
        index["total_assessments"] = sum(
            len(u["assessments"]) for u in index["users"].values()
        )

        with open(index_file, 'w', encoding='utf-8') as f:
            json.dump(index, f, ensure_ascii=False, indent=2)

    def get_assessment(self, user_id: str, date: str = None) -> Optional[Dict]:
        """
        获取用户评估数据

        Args:
            user_id: 用户ID
            date: 评估日期 (默认获取最新)

        Returns:
            评估数据字典
        """
        user_dir = self.processed_dir / "users" / user_id / "assessments"
        if not user_dir.exists():
            return None

        if date:
            ass_file = user_dir / f"{date}.json"
        else:
            # 获取最新
            files = sorted(user_dir.glob("*.json"), reverse=True)
            if not files:
                return None
            ass_file = files[0]

        if ass_file.exists():
            with open(ass_file, 'r', encoding='utf-8') as f:
                return json.load(f)

        return None

    def list_users(self) -> List[Dict]:
        """列出所有用户"""
        index_file = self.processed_dir / "index.json"
        if not index_file.exists():
            return []

        with open(index_file, 'r', encoding='utf-8') as f:
            index = json.load(f)

        return [
            {"user_id": uid, **info}
            for uid, info in index.get("users", {}).items()
        ]


# ============ 主函数 ============

def main():
    parser = argparse.ArgumentParser(description="评估数据导入器")
    parser.add_argument("--source", required=True, help="源数据目录")
    parser.add_argument("--batch", default=None, help="批次名称")
    parser.add_argument("--no-copy", action="store_true", help="不复制原始文件")
    parser.add_argument("--generate-prescriptions", action="store_true", help="自动生成处方")

    args = parser.parse_args()

    importer = AssessmentImporter()
    assessments = importer.import_from_directory(
        source_dir=args.source,
        batch_name=args.batch,
        copy_files=not args.no_copy
    )

    # 生成处方
    if args.generate_prescriptions and assessments:
        print(f"\n正在生成处方...")
        try:
            from prescription_engine import generate_prescription

            for ass in assessments:
                physio = {
                    "SDNN": ass.physio_data.hrv.sdnn,
                    "RMSSD": ass.physio_data.hrv.rmssd,
                    "heart_rate": ass.physio_data.heart_rate.avg
                }
                psych = {
                    "fatigue_index": ass.psych_data.fatigue_index,
                    "energy": 100 - ass.psych_data.fatigue_index,
                    "mood": ass.psych_data.mood_index,
                    "stress_index": ass.psych_data.stress_index
                }

                prescription = generate_prescription(physio, psych)

                # 保存处方
                rx_dir = importer.processed_dir / "users" / ass.user_id / "prescriptions"
                rx_dir.mkdir(parents=True, exist_ok=True)
                rx_file = rx_dir / f"{ass.assessment_date}.json"

                with open(rx_file, 'w', encoding='utf-8') as f:
                    json.dump(prescription, f, ensure_ascii=False, indent=2)

                print(f"  [处方] {ass.user_id}: {prescription['prescription_meta']['name']}")

        except ImportError:
            print("[警告] 无法导入 prescription_engine，跳过处方生成")

    print(f"\n完成！数据已保存到: {importer.data_dir}")


if __name__ == "__main__":
    main()
