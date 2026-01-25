#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
data_extractor.py - 健康数据提取模块

优先级策略：
1. 优先从 Excel (.xlsx) 读取生理指标（SDNN, RMSSD, HRV 等）
2. PDF 仅用于提取焦虑分数和 SCL-90 因子分
3. 使用正则表达式模糊匹配提取数值
4. 异常回退机制：解析失败时给出明确提示
"""

import re
import sys
import io
import os
from pathlib import Path
from typing import Dict, Any, Optional, List, Tuple
from dataclasses import dataclass, field
from datetime import datetime

import pandas as pd

# Windows 控制台编码修复（仅在直接运行时执行，避免重复包装）
def _fix_windows_encoding():
    """修复 Windows 控制台编码问题"""
    if sys.platform == 'win32':
        try:
            # 检查是否已经是 TextIOWrapper
            if not isinstance(sys.stdout, io.TextIOWrapper) or sys.stdout.encoding != 'utf-8':
                if hasattr(sys.stdout, 'buffer'):
                    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
            if not isinstance(sys.stderr, io.TextIOWrapper) or sys.stderr.encoding != 'utf-8':
                if hasattr(sys.stderr, 'buffer'):
                    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
        except Exception:
            pass  # 忽略编码修复失败


@dataclass
class ExtractionResult:
    """数据提取结果"""
    # 生理数据（来自 Excel）
    physio_data: Dict[str, Any] = field(default_factory=dict)
    physio_source: str = ""  # 数据来源文件
    physio_success: bool = False

    # 心理数据（来自 PDF 或 Excel）
    psych_data: Dict[str, Any] = field(default_factory=dict)
    psych_source: str = ""
    psych_success: bool = False

    # 解析失败的字段
    failed_fields: List[str] = field(default_factory=list)

    # 警告信息
    warnings: List[str] = field(default_factory=list)

    # 原始数据（用于调试）
    raw_excel_physio: Optional[pd.DataFrame] = None
    raw_excel_psych: Optional[pd.DataFrame] = None
    raw_pdf_text: str = ""


# ============ 正则表达式模式 ============
# 增强模糊匹配：使用 [\s:：=\-_\(\)（）\[\]【】]* 容忍任意空格和特殊字符
# SEP 用于匹配关键词和数值之间的任意分隔符
SEP = r'[\s:：=\-_\(\)（）\[\]【】\|｜/／\.,，。]*'

REGEX_PATTERNS = {
    # HRV 生理指标（通常在 Excel 中）
    "SDNN": rf"SDNN{SEP}(\d+\.?\d*)",
    "RMSSD": rf"RMSSD{SEP}(\d+\.?\d*)",
    "pNN50": rf"[pP]NN50{SEP}(\d+\.?\d*)",
    "LF": rf"(?<![A-Za-z])LF{SEP}(\d+\.?\d*)",
    "HF": rf"(?<![A-Za-z])HF{SEP}(\d+\.?\d*)",
    "LF_HF_ratio": rf"LF[/／\s]*HF{SEP}(\d+\.?\d*)",
    "TP": rf"(?:TP|总功率|Total\s*Power){SEP}(\d+\.?\d*)",

    # 心率指标
    "heart_rate": rf"(?:平均心率|心率|HR|Heart\s*Rate){SEP}(\d+\.?\d*)",
    "heart_rate_max": rf"(?:最大心率|最高心率|Max\s*HR|峰值心率){SEP}(\d+\.?\d*)",
    "heart_rate_min": rf"(?:最小心率|最低心率|Min\s*HR|谷值心率){SEP}(\d+\.?\d*)",

    # 心理指标（需要从 PDF 提取）
    "anxiety_score": rf"(?:焦虑[分得评指]?[数分值]?|SAS[分得]?[数分]?|焦虑量表|焦虑自评|Anxiety){SEP}(\d+\.?\d*)",
    "depression_score": rf"(?:抑郁[分得评指]?[数分值]?|SDS[分得]?[数分]?|抑郁量表|抑郁自评|Depression){SEP}(\d+\.?\d*)",
    "stress_index": rf"(?:压力[指分]?[数分值]?|心理压力|Stress){SEP}(\d+\.?\d*)",
    "fatigue_index": rf"(?:疲劳[指分]?[数分值]?|疲劳度|Fatigue){SEP}(\d+\.?\d*)",

    # SCL-90 因子分 - 增强模糊匹配
    "scl90_total": rf"(?:SCL[-_\s]?90|症状自评){SEP}(?:总分|总均分|总|Total)?{SEP}(\d+\.?\d*)",
    "scl90_somatization": rf"(?:躯体化|躯体不适|Somatization){SEP}(\d+\.?\d*)",
    "scl90_obsessive": rf"(?:强迫[症状]?|强迫观念|Obsessive){SEP}(\d+\.?\d*)",
    "scl90_interpersonal": rf"(?:人际[关系敏感]?|人际关系|Interpersonal){SEP}(\d+\.?\d*)",
    "scl90_depression": rf"(?:抑郁[症状因子]?|SCL.*抑郁){SEP}(\d+\.?\d*)",
    "scl90_anxiety": rf"(?:焦虑[症状因子]?|SCL.*焦虑){SEP}(\d+\.?\d*)",
    "scl90_hostility": rf"(?:敌对[性情]?|敌意|Hostility){SEP}(\d+\.?\d*)",
    "scl90_phobic": rf"(?:恐怖|恐惧|Phobic){SEP}(\d+\.?\d*)",
    "scl90_paranoid": rf"(?:偏执|妄想|Paranoid){SEP}(\d+\.?\d*)",
    "scl90_psychoticism": rf"(?:精神病性|精神质|Psychoticism){SEP}(\d+\.?\d*)",
    "scl90_other": rf"(?:其他|附加[项因子]?|睡眠饮食){SEP}(\d+\.?\d*)",
}

# Excel 列名映射（中文 -> 标准名）
EXCEL_COLUMN_MAP = {
    # 生理指标
    "SDNN": "SDNN",
    "RMSSD": "RMSSD",
    "PNN50": "pNN50",
    "pNN50": "pNN50",
    "LF": "LF",
    "HF": "HF",
    "平均心率": "heart_rate",
    "最大心率": "heart_rate_max",
    "最小心率": "heart_rate_min",
    "最高心率": "heart_rate_max",
    "最低心率": "heart_rate_min",
    "心率标准差": "heart_rate_std",
    "变异系数": "cv",
    "RRI均值": "rri_mean",
    "呼吸": "respiration",
    "交感状态": "sympathetic_state",
    "低频归一化单位": "lf_norm",
    "高频归一化单位": "hf_norm",

    # 心理指标
    "压力": "stress_index",
    "疲劳": "fatigue_index",
    "精力": "energy",
    "情绪": "mood",
    "情绪状态": "mood_state",
}


def find_matching_files(data_dir: str, user_id: str = None) -> Dict[str, List[Path]]:
    """
    在数据目录中查找匹配的 Excel 和 PDF 文件

    Args:
        data_dir: 数据目录路径
        user_id: 可选的用户 ID（如 C302A07E7F52）

    Returns:
        包含 'physio_excel', 'psych_excel', 'pdf' 的文件路径字典
    """
    data_path = Path(data_dir)
    result = {
        "physio_excel": [],
        "psych_excel": [],
        "pdf": []
    }

    if not data_path.exists():
        return result

    for file in data_path.iterdir():
        if not file.is_file():
            continue

        name = file.name

        # 如果指定了 user_id，只匹配对应用户的文件
        if user_id and user_id not in name:
            continue

        if name.endswith('.xlsx'):
            if '生理测评' in name:
                result["physio_excel"].append(file)
            elif '心理测评' in name:
                result["psych_excel"].append(file)
        elif name.endswith('.pdf'):
            if '心理健康' in name or '测评报告' in name:
                result["pdf"].append(file)

    # 按修改时间排序，最新的在前
    for key in result:
        result[key].sort(key=lambda x: x.stat().st_mtime, reverse=True)

    return result


def _clean_column_name(col_name: str) -> str:
    """
    清洗列名：去掉空格、换行符、特殊字符

    Args:
        col_name: 原始列名

    Returns:
        清洗后的列名
    """
    if not isinstance(col_name, str):
        col_name = str(col_name)
    # 去掉空格、换行符、制表符
    cleaned = re.sub(r'[\s\n\r\t]+', '', col_name)
    # 去掉特殊字符但保留中文和字母数字
    cleaned = re.sub(r'[^\w\u4e00-\u9fff]', '', cleaned)
    return cleaned


def _get_sheet_count(excel_path: str) -> int:
    """
    使用 openpyxl 获取 Sheet 数量

    Args:
        excel_path: Excel 文件路径

    Returns:
        Sheet 数量
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        count = len(wb.sheetnames)
        wb.close()
        return count
    except Exception:
        return 1


def _find_best_sheet_index(excel_path: str) -> Tuple[int, List[str]]:
    """
    使用 openpyxl 查找最佳 Sheet 索引

    优先级匹配关键词，返回索引而非名称

    Args:
        excel_path: Excel 文件路径

    Returns:
        (最佳 Sheet 索引, 所有 Sheet 名列表)
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        if not sheet_names:
            return 0, []

        # 优先级匹配关键词
        priority_keywords = [
            ['生理', '测评'],
            ['生理'],
            ['HRV'],
            ['physio'],
            ['心率', '变异'],
            ['数据'],
        ]

        for keywords in priority_keywords:
            for idx, sheet in enumerate(sheet_names):
                sheet_lower = sheet.lower()
                if all(kw.lower() in sheet_lower for kw in keywords):
                    return idx, sheet_names

        # 如果没有匹配，返回第一个 Sheet（索引 0）
        return 0, sheet_names

    except Exception as e:
        print(f"[警告] 读取 Sheet 列表失败: {e}")
        return 0, []


def extract_from_excel_physio(excel_path: str) -> Tuple[Dict[str, Any], pd.DataFrame]:
    """
    从生理测评 Excel 文件中提取数据

    改进：
    1. 使用 openpyxl 作为引擎
    2. 通过 Sheet 索引加载数据（避免编码问题）
    3. 清洗列名（去掉空格和换行符）

    Args:
        excel_path: Excel 文件路径

    Returns:
        (提取的数据字典, 原始 DataFrame)
    """
    try:
        # 使用 openpyxl 查找最佳 Sheet 索引
        target_idx, all_sheets = _find_best_sheet_index(excel_path)

        if not all_sheets:
            print(f"[警告] Excel 文件中没有找到任何 Sheet")
            return {}, pd.DataFrame()

        print(f"  [Sheet 匹配] 找到 {len(all_sheets)} 个 Sheet: {all_sheets}")
        print(f"  [Sheet 匹配] 选择读取索引 {target_idx}: '{all_sheets[target_idx]}'")

        # 使用 openpyxl 引擎，通过索引加载
        df = pd.read_excel(
            excel_path,
            sheet_name=target_idx,  # 使用索引而非名称
            engine='openpyxl'
        )

        if df.empty:
            return {}, df

        # 清洗列名：去掉空格、换行符
        original_columns = list(df.columns)
        cleaned_columns = [_clean_column_name(col) for col in df.columns]
        df.columns = cleaned_columns

        # 打印列名映射（调试用）
        print(f"  [列名清洗] 原始 -> 清洗后:")
        for orig, clean in zip(original_columns, cleaned_columns):
            if orig != clean:
                print(f"    '{orig}' -> '{clean}'")

        # 过滤掉全零行，获取有效数据
        numeric_cols = df.select_dtypes(include=['number']).columns
        if len(numeric_cols) == 0:
            print(f"[警告] Sheet 索引 {target_idx} 中没有数值列")
            return {}, df

        df_valid = df[df[numeric_cols].sum(axis=1) > 0]

        if df_valid.empty:
            # 尝试读取其他 Sheet（通过索引）
            for other_idx in range(len(all_sheets)):
                if other_idx == target_idx:
                    continue
                print(f"  [回退] 尝试读取 Sheet 索引 {other_idx}: '{all_sheets[other_idx]}'")
                df_other = pd.read_excel(excel_path, sheet_name=other_idx, engine='openpyxl')
                # 清洗列名
                df_other.columns = [_clean_column_name(col) for col in df_other.columns]
                numeric_cols_other = df_other.select_dtypes(include=['number']).columns
                if len(numeric_cols_other) > 0:
                    df_valid_other = df_other[df_other[numeric_cols_other].sum(axis=1) > 0]
                    if not df_valid_other.empty:
                        df = df_other
                        df_valid = df_valid_other
                        print(f"  [回退] 成功从索引 {other_idx} 读取到数据")
                        break

        if df_valid.empty:
            return {}, df

        # 取最后 5 行的中位数
        last_rows = df_valid.tail(5)

        extracted = {}

        # 构建清洗后的列名映射
        cleaned_column_map = {_clean_column_name(k): v for k, v in EXCEL_COLUMN_MAP.items()}

        for clean_col in df.columns:
            if clean_col in cleaned_column_map:
                std_name = cleaned_column_map[clean_col]
                values = last_rows[clean_col].dropna()
                if len(values) > 0:
                    if values.dtype in ['int64', 'float64']:
                        extracted[std_name] = round(float(values.median()), 2)
                    else:
                        extracted[std_name] = str(values.iloc[-1])

        return extracted, df

    except Exception as e:
        print(f"[错误] 读取生理测评 Excel 失败: {e}")
        return {}, pd.DataFrame()


def extract_from_excel_psych(excel_path: str) -> Tuple[Dict[str, Any], pd.DataFrame]:
    """
    从心理测评 Excel 文件中提取数据

    Args:
        excel_path: Excel 文件路径

    Returns:
        (提取的数据字典, 原始 DataFrame)
    """
    try:
        df = pd.read_excel(excel_path)

        if df.empty:
            return {}, df

        # 取最后一行数据
        last_row = df.iloc[-1]

        extracted = {}

        for excel_col, std_name in EXCEL_COLUMN_MAP.items():
            if excel_col in df.columns:
                value = last_row[excel_col]
                if pd.notna(value):
                    if isinstance(value, (int, float)):
                        extracted[std_name] = round(float(value), 2)
                    elif std_name in ['mood_state']:  # 保留文本字段
                        extracted[std_name] = str(value)
                    else:
                        # 尝试转换为数值
                        try:
                            extracted[std_name] = round(float(value), 2)
                        except (ValueError, TypeError):
                            extracted[std_name] = str(value)

        return extracted, df

    except Exception as e:
        print(f"[错误] 读取心理测评 Excel 失败: {e}")
        return {}, pd.DataFrame()


def _extract_value_after_anchor(text: str, anchor: str, search_range: int = 200) -> Optional[float]:
    """
    关键词锚点法：定位到关键词后，向后搜索第一个浮点数

    Args:
        text: 全文本
        anchor: 锚点关键词
        search_range: 向后搜索的字符范围

    Returns:
        找到的浮点数，或 None
    """
    # 查找锚点位置
    anchor_pos = text.find(anchor)
    if anchor_pos == -1:
        # 尝试模糊匹配（忽略空格）
        anchor_no_space = anchor.replace(' ', '')
        text_no_space = text.replace(' ', '').replace('\n', '')
        anchor_pos = text_no_space.find(anchor_no_space)
        if anchor_pos == -1:
            return None
        # 在原文中找到对应位置（近似）
        anchor_pos = text.find(anchor[0], max(0, anchor_pos - 10))

    if anchor_pos == -1:
        return None

    # 向后搜索第一个浮点数
    search_text = text[anchor_pos:anchor_pos + search_range]
    # 匹配整数或浮点数
    match = re.search(r'(\d+\.?\d*)', search_text)
    if match:
        try:
            return float(match.group(1))
        except ValueError:
            return None
    return None


def extract_from_pdf_with_regex(pdf_path: str) -> Tuple[Dict[str, Any], str]:
    """
    从 PDF 中使用关键词锚点法提取焦虑分数和 SCL-90 因子分

    改进：
    1. 不使用 extract_table，改用全文本搜索
    2. 先定位关键词锚点，向后搜索第一个浮点数
    3. 如果 raw_text 为空，明确提示是扫描件

    Args:
        pdf_path: PDF 文件路径

    Returns:
        (提取的数据字典, 原始文本)
    """
    raw_text = ""
    extracted = {}

    print(f"  [PDF 解析] 开始全文本提取（不使用 extract_table）...")

    # 尝试方法1：pdfplumber 直接提取文本（不使用 extract_table）
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            for i, page in enumerate(pdf.pages):
                # 只提取文本，不提取表格
                text = page.extract_text()
                if text:
                    raw_text += text + "\n"
                    print(f"    第 {i+1} 页: 提取到 {len(text)} 字符")
    except Exception as e:
        print(f"  [警告] pdfplumber 提取失败: {e}")

    # 如果 pdfplumber 没有提取到文本，尝试 PyMuPDF
    if not raw_text.strip():
        print(f"  [回退] pdfplumber 无文本，尝试 PyMuPDF...")
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(pdf_path)
            for i, page in enumerate(doc):
                text = page.get_text()
                if text:
                    raw_text += text + "\n"
                    print(f"    第 {i+1} 页: 提取到 {len(text)} 字符")
            doc.close()
        except Exception as e:
            print(f"  [警告] PyMuPDF 提取失败: {e}")

    # 如果仍然没有文本，明确提示是扫描件
    if not raw_text.strip():
        print(f"  [警告] ⚠️ PDF 无文本层，请确认是否为扫描件")
        print(f"  [提示] 扫描件需要 OCR 支持，请安装 pytesseract 和 pdf2image")

        # 尝试 OCR（如果可用）
        try:
            raw_text = _extract_text_with_ocr(pdf_path)
            if raw_text.strip():
                print(f"  [OCR] 成功提取 {len(raw_text)} 字符")
        except Exception as e:
            print(f"  [警告] OCR 提取失败: {e}")

        if not raw_text.strip():
            return {}, ""

    # 使用关键词锚点法提取数据
    print(f"  [PDF 解析] 使用关键词锚点法搜索数据...")

    # 关键词锚点配置：(标准名, [锚点关键词列表])
    anchor_config = [
        # 心理健康报告锚点
        ("report_title", ["心理健康测评个人报告", "心理健康测评报告", "个人报告"]),

        # 焦虑相关
        ("anxiety_score", ["焦虑分数", "焦虑得分", "焦虑评分", "SAS分数", "SAS得分", "焦虑自评", "焦虑"]),

        # 抑郁相关
        ("depression_score", ["抑郁分数", "抑郁得分", "抑郁评分", "SDS分数", "SDS得分", "抑郁自评", "抑郁"]),

        # 压力相关
        ("stress_index", ["压力指数", "压力分数", "压力值", "心理压力"]),

        # SCL-90 因子分
        ("scl90_total", ["SCL-90总分", "SCL90总分", "症状自评总分", "总均分"]),
        ("scl90_somatization", ["躯体化", "躯体不适"]),
        ("scl90_obsessive", ["强迫症状", "强迫"]),
        ("scl90_interpersonal", ["人际关系敏感", "人际敏感", "人际关系"]),
        ("scl90_depression", ["抑郁因子", "抑郁症状"]),
        ("scl90_anxiety", ["焦虑因子", "焦虑症状"]),
        ("scl90_hostility", ["敌对", "敌意"]),
        ("scl90_phobic", ["恐怖", "恐惧"]),
        ("scl90_paranoid", ["偏执", "妄想"]),
        ("scl90_psychoticism", ["精神病性", "精神质"]),
    ]

    for std_name, anchors in anchor_config:
        if std_name == "report_title":
            # 只检查报告标题是否存在
            for anchor in anchors:
                if anchor in raw_text:
                    print(f"    [锚点] 找到报告标题: '{anchor}'")
                    break
            continue

        # 尝试每个锚点
        for anchor in anchors:
            value = _extract_value_after_anchor(raw_text, anchor)
            if value is not None:
                extracted[std_name] = value
                print(f"    [锚点] {std_name}: 通过 '{anchor}' 找到值 {value}")
                break

    # 如果锚点法没找到，回退到正则匹配
    if not extracted:
        print(f"  [回退] 锚点法未匹配，尝试正则表达式...")
        psych_patterns = {
            "anxiety_score": REGEX_PATTERNS["anxiety_score"],
            "depression_score": REGEX_PATTERNS["depression_score"],
            "stress_index": REGEX_PATTERNS["stress_index"],
            "scl90_total": REGEX_PATTERNS["scl90_total"],
            "scl90_somatization": REGEX_PATTERNS["scl90_somatization"],
            "scl90_obsessive": REGEX_PATTERNS["scl90_obsessive"],
            "scl90_interpersonal": REGEX_PATTERNS["scl90_interpersonal"],
            "scl90_depression": REGEX_PATTERNS["scl90_depression"],
            "scl90_anxiety": REGEX_PATTERNS["scl90_anxiety"],
            "scl90_hostility": REGEX_PATTERNS["scl90_hostility"],
            "scl90_phobic": REGEX_PATTERNS["scl90_phobic"],
            "scl90_paranoid": REGEX_PATTERNS["scl90_paranoid"],
            "scl90_psychoticism": REGEX_PATTERNS["scl90_psychoticism"],
        }

        for key, pattern in psych_patterns.items():
            if key not in extracted:
                match = re.search(pattern, raw_text, re.IGNORECASE)
                if match:
                    try:
                        value = float(match.group(1))
                        extracted[key] = value
                        print(f"    [正则] {key}: {value}")
                    except (ValueError, IndexError):
                        pass

    if not extracted:
        print(f"  [结果] 未能从 PDF 中提取到心理数据")
    else:
        print(f"  [结果] 成功提取 {len(extracted)} 项心理指标")

    return extracted, raw_text


def _extract_text_with_ocr(pdf_path: str) -> str:
    """
    使用 OCR 从扫描件 PDF 提取文本

    需要安装: pip install pytesseract pdf2image
    以及系统安装 Tesseract-OCR
    """
    try:
        from pdf2image import convert_from_path
        import pytesseract

        # 转换 PDF 为图片
        images = convert_from_path(pdf_path, dpi=200)

        all_text = []
        for i, image in enumerate(images):
            # 使用中文+英文识别
            text = pytesseract.image_to_string(image, lang='chi_sim+eng')
            all_text.append(text)

        return "\n".join(all_text)

    except ImportError:
        print("[提示] OCR 需要安装 pytesseract 和 pdf2image")
        print("       pip install pytesseract pdf2image")
        print("       同时需要系统安装 Tesseract-OCR")
        return ""
    except Exception as e:
        print(f"[错误] OCR 处理失败: {e}")
        return ""


def extract_health_data(
    data_dir: str,
    user_id: str = None,
    prefer_latest: bool = True
) -> ExtractionResult:
    """
    提取健康数据的主函数

    优先级策略：
    1. 优先从 Excel 读取生理指标
    2. PDF 仅用于提取焦虑分数和 SCL-90 因子分
    3. 异常回退：解析失败时给出明确提示

    Args:
        data_dir: 数据目录路径（如 data/raw/）
        user_id: 可选的用户 ID 筛选
        prefer_latest: 是否优先使用最新文件

    Returns:
        ExtractionResult 对象
    """
    result = ExtractionResult()

    # 1. 查找匹配的文件
    files = find_matching_files(data_dir, user_id)

    print(f"\n{'='*60}")
    print("健康数据提取开始")
    print(f"{'='*60}")
    print(f"数据目录: {data_dir}")
    print(f"用户 ID: {user_id or '全部'}")
    print(f"找到的文件:")
    print(f"  - 生理测评 Excel: {len(files['physio_excel'])} 个")
    print(f"  - 心理测评 Excel: {len(files['psych_excel'])} 个")
    print(f"  - PDF 报告: {len(files['pdf'])} 个")

    # 2. 优先从 Excel 读取生理数据
    if files['physio_excel']:
        physio_file = files['physio_excel'][0]  # 最新的文件
        print(f"\n[生理数据] 读取 Excel: {physio_file.name}")

        physio_data, raw_df = extract_from_excel_physio(str(physio_file))

        if physio_data:
            result.physio_data = physio_data
            result.physio_source = str(physio_file)
            result.physio_success = True
            result.raw_excel_physio = raw_df
            print(f"  -> 成功提取 {len(physio_data)} 项生理指标")
            for k, v in physio_data.items():
                print(f"     {k}: {v}")
        else:
            result.warnings.append("生理测评 Excel 文件读取失败或无有效数据")
            result.failed_fields.append("physio_excel")
    else:
        result.warnings.append("未找到生理测评 Excel 文件")
        result.failed_fields.append("physio_excel")

    # 3. 从心理测评 Excel 读取基础心理数据
    if files['psych_excel']:
        psych_file = files['psych_excel'][0]
        print(f"\n[心理数据] 读取 Excel: {psych_file.name}")

        psych_excel_data, raw_df = extract_from_excel_psych(str(psych_file))

        if psych_excel_data:
            result.psych_data.update(psych_excel_data)
            result.psych_source = str(psych_file)
            result.raw_excel_psych = raw_df
            print(f"  -> 成功提取 {len(psych_excel_data)} 项心理指标")
            for k, v in psych_excel_data.items():
                print(f"     {k}: {v}")

    # 4. 从 PDF 提取焦虑分数和 SCL-90（作为补充）
    if files['pdf']:
        pdf_file = files['pdf'][0]
        print(f"\n[PDF 解析] 尝试提取焦虑分数和 SCL-90: {pdf_file.name}")

        pdf_data, raw_text = extract_from_pdf_with_regex(str(pdf_file))
        result.raw_pdf_text = raw_text

        if pdf_data:
            # PDF 数据作为补充，不覆盖已有数据
            for k, v in pdf_data.items():
                if k not in result.psych_data:
                    result.psych_data[k] = v
            result.psych_success = True
            print(f"  -> 成功提取 {len(pdf_data)} 项心理指标")
            for k, v in pdf_data.items():
                print(f"     {k}: {v}")
        else:
            if not raw_text.strip():
                msg = "PDF 为扫描件格式，文本提取失败（需要 OCR 支持）"
            else:
                msg = "PDF 文本中未匹配到焦虑分数或 SCL-90 因子分"
            result.warnings.append(msg)
            print(f"  -> 警告: {msg}")

    # 5. 检查心理数据是否成功
    if result.psych_data:
        result.psych_success = True
    else:
        result.failed_fields.append("psych_data")

    # 6. 生成异常回退提示
    print(f"\n{'='*60}")
    print("提取结果汇总")
    print(f"{'='*60}")
    print(f"生理数据: {'成功' if result.physio_success else '失败'} (来源: {result.physio_source or '无'})")
    print(f"心理数据: {'成功' if result.psych_success else '失败'} (来源: {result.psych_source or '无'})")

    if result.warnings:
        print(f"\n警告信息:")
        for w in result.warnings:
            print(f"  - {w}")

    if result.failed_fields:
        print(f"\n解析失败提示:")
        if "physio_excel" in result.failed_fields:
            print("  - 生理数据提取失败，请手动确认")
        if "psych_data" in result.failed_fields:
            print("  - 心理数据解析失败，请手动确认")

    return result


def generate_fallback_message(result: ExtractionResult) -> str:
    """
    生成异常回退提示信息

    Args:
        result: 提取结果

    Returns:
        提示信息字符串
    """
    messages = []

    if result.physio_success:
        messages.append(f"生理数据提取自 Excel ({Path(result.physio_source).name})")
    else:
        messages.append("生理数据提取失败，请手动确认")

    if result.psych_success:
        if result.psych_source:
            messages.append(f"心理数据提取自 {Path(result.psych_source).name}")
    else:
        messages.append("心理数据解析失败，请手动确认")

    if result.warnings:
        messages.append(f"注意事项: {'; '.join(result.warnings)}")

    return "\n".join(messages)


# ============ 看板降级策略 ============

@dataclass
class DashboardData:
    """看板数据结构"""
    # 看板类型：'full' 完整看板, 'partial' 初步看板（仅生理数据）
    dashboard_type: str = "full"

    # 生理指标看板
    hrv_metrics: Dict[str, Any] = field(default_factory=dict)
    heart_rate_metrics: Dict[str, Any] = field(default_factory=dict)

    # 心理指标看板（可能为空）
    psych_metrics: Dict[str, Any] = field(default_factory=dict)

    # 综合评估
    overall_status: str = ""
    risk_level: str = ""  # 'low', 'medium', 'high'

    # 警告和提示
    notices: List[str] = field(default_factory=list)

    # 数据来源
    data_sources: List[str] = field(default_factory=list)


def _assess_hrv_status(physio_data: Dict[str, Any]) -> Tuple[str, str]:
    """
    根据 HRV 指标评估健康状态

    Returns:
        (状态描述, 风险等级)
    """
    sdnn = physio_data.get('SDNN', 0)
    rmssd = physio_data.get('RMSSD', 0)
    lf_hf = physio_data.get('LF_HF_ratio', 1.0)

    # 评估逻辑
    if sdnn >= 100 and rmssd >= 40:
        return "HRV 指标良好，自主神经调节功能正常", "low"
    elif sdnn >= 50 and rmssd >= 20:
        return "HRV 指标中等，建议适当放松", "medium"
    else:
        return "HRV 指标偏低，建议关注压力管理", "high"


def _assess_heart_rate_status(physio_data: Dict[str, Any]) -> str:
    """根据心率指标评估状态"""
    hr = physio_data.get('heart_rate', 0)
    hr_max = physio_data.get('heart_rate_max', 0)
    hr_min = physio_data.get('heart_rate_min', 0)

    if 60 <= hr <= 100:
        return f"心率正常 ({hr} bpm)"
    elif hr < 60:
        return f"心率偏低 ({hr} bpm)，可能为运动员心脏或需关注"
    else:
        return f"心率偏高 ({hr} bpm)，建议放松休息"


def generate_dashboard(result: ExtractionResult) -> DashboardData:
    """
    生成健康看板

    降级策略：
    - 如果生理和心理数据都有 -> 生成完整看板
    - 如果只有生理数据 -> 生成初步看板，标注心理数据缺失
    - 如果生理数据也没有 -> 返回空看板，提示数据缺失

    Args:
        result: 数据提取结果

    Returns:
        DashboardData 看板数据
    """
    dashboard = DashboardData()

    # 情况 1：完全没有数据
    if not result.physio_success and not result.psych_success:
        dashboard.dashboard_type = "empty"
        dashboard.overall_status = "数据提取失败"
        dashboard.risk_level = "unknown"
        dashboard.notices.append("生理数据和心理数据均提取失败，请手动确认原始文件")
        return dashboard

    # 情况 2：只有生理数据（降级模式）
    if result.physio_success and not result.psych_success:
        dashboard.dashboard_type = "partial"
        dashboard.notices.append("⚠️ 心理数据缺失，请参考 PDF 报告明细")
        dashboard.notices.append("当前为【初步看板】，仅展示生理指标")

    # 情况 3：完整数据
    elif result.physio_success and result.psych_success:
        dashboard.dashboard_type = "full"

    # 填充生理数据
    if result.physio_success:
        physio = result.physio_data

        # HRV 指标
        dashboard.hrv_metrics = {
            "SDNN": physio.get('SDNN', '--'),
            "RMSSD": physio.get('RMSSD', '--'),
            "pNN50": physio.get('pNN50', '--'),
            "LF": physio.get('LF', '--'),
            "HF": physio.get('HF', '--'),
            "LF/HF": physio.get('LF_HF_ratio', '--'),
        }

        # 心率指标
        dashboard.heart_rate_metrics = {
            "平均心率": physio.get('heart_rate', '--'),
            "最高心率": physio.get('heart_rate_max', '--'),
            "最低心率": physio.get('heart_rate_min', '--'),
            "心率标准差": physio.get('heart_rate_std', '--'),
        }

        # 评估状态
        hrv_status, risk = _assess_hrv_status(physio)
        hr_status = _assess_heart_rate_status(physio)

        dashboard.overall_status = f"{hrv_status}；{hr_status}"
        dashboard.risk_level = risk

        dashboard.data_sources.append(f"生理数据: {Path(result.physio_source).name}")

    # 填充心理数据
    if result.psych_success:
        psych = result.psych_data

        dashboard.psych_metrics = {
            "压力指数": psych.get('stress_index', '--'),
            "疲劳指数": psych.get('fatigue_index', '--'),
            "精力值": psych.get('energy', '--'),
            "情绪值": psych.get('mood', '--'),
            "情绪状态": psych.get('mood_state', '--'),
            "焦虑评分": psych.get('anxiety_score', '--'),
            "抑郁评分": psych.get('depression_score', '--'),
        }

        # SCL-90 因子分（如果有）
        scl90_keys = [k for k in psych.keys() if k.startswith('scl90_')]
        if scl90_keys:
            dashboard.psych_metrics["SCL-90 因子"] = {
                k.replace('scl90_', ''): psych[k] for k in scl90_keys
            }

        if result.psych_source:
            dashboard.data_sources.append(f"心理数据: {Path(result.psych_source).name}")

    return dashboard


def format_dashboard_text(dashboard: DashboardData) -> str:
    """
    将看板数据格式化为文本输出

    Args:
        dashboard: 看板数据

    Returns:
        格式化的文本字符串
    """
    lines = []

    # 标题
    if dashboard.dashboard_type == "full":
        lines.append("=" * 60)
        lines.append("        健康状态看板 - 完整版")
        lines.append("=" * 60)
    elif dashboard.dashboard_type == "partial":
        lines.append("=" * 60)
        lines.append("        健康状态看板 - 初步版（仅生理数据）")
        lines.append("=" * 60)
    else:
        lines.append("=" * 60)
        lines.append("        健康状态看板 - 数据缺失")
        lines.append("=" * 60)

    # 警告信息
    if dashboard.notices:
        lines.append("")
        for notice in dashboard.notices:
            lines.append(f"  {notice}")
        lines.append("")

    # 综合评估
    if dashboard.overall_status:
        lines.append("-" * 40)
        lines.append("【综合评估】")
        lines.append(f"  {dashboard.overall_status}")
        risk_emoji = {"low": "🟢", "medium": "🟡", "high": "🔴", "unknown": "⚪"}
        lines.append(f"  风险等级: {risk_emoji.get(dashboard.risk_level, '⚪')} {dashboard.risk_level.upper()}")
        lines.append("")

    # HRV 指标
    if dashboard.hrv_metrics:
        lines.append("-" * 40)
        lines.append("【HRV 心率变异性指标】")
        for k, v in dashboard.hrv_metrics.items():
            unit = "ms" if k in ["SDNN", "RMSSD"] else ("%" if k == "pNN50" else "ms²" if k in ["LF", "HF"] else "")
            lines.append(f"  {k}: {v} {unit}".strip())
        lines.append("")

    # 心率指标
    if dashboard.heart_rate_metrics:
        lines.append("-" * 40)
        lines.append("【心率指标】")
        for k, v in dashboard.heart_rate_metrics.items():
            unit = "bpm" if "心率" in k else ""
            lines.append(f"  {k}: {v} {unit}".strip())
        lines.append("")

    # 心理指标
    if dashboard.psych_metrics:
        lines.append("-" * 40)
        lines.append("【心理状态指标】")
        for k, v in dashboard.psych_metrics.items():
            if k == "SCL-90 因子" and isinstance(v, dict):
                lines.append(f"  {k}:")
                for sk, sv in v.items():
                    lines.append(f"    - {sk}: {sv}")
            else:
                lines.append(f"  {k}: {v}")
        lines.append("")

    # 数据来源
    if dashboard.data_sources:
        lines.append("-" * 40)
        lines.append("【数据来源】")
        for src in dashboard.data_sources:
            lines.append(f"  • {src}")

    lines.append("=" * 60)

    return "\n".join(lines)


def extract_and_generate_dashboard(
    data_dir: str,
    user_id: str = None,
    print_output: bool = True
) -> Tuple[ExtractionResult, DashboardData]:
    """
    一站式提取数据并生成看板

    降级策略自动处理：
    - PDF 提取失败时，仅根据 Excel 生理指标生成初步看板
    - 不会因为心理数据缺失而中断流程

    Args:
        data_dir: 数据目录
        user_id: 可选用户 ID
        print_output: 是否打印看板

    Returns:
        (提取结果, 看板数据)
    """
    # 提取数据
    result = extract_health_data(data_dir, user_id)

    # 生成看板
    dashboard = generate_dashboard(result)

    # 打印看板
    if print_output:
        print("\n" + format_dashboard_text(dashboard))

    return result, dashboard


# ============ 主函数入口 ============
if __name__ == "__main__":
    import argparse

    # 仅在直接运行时修复编码
    _fix_windows_encoding()

    parser = argparse.ArgumentParser(description="健康数据提取与看板生成")
    parser.add_argument("--data-dir", default=r"J:\xingjian-agent\MyOctopusProject\data\raw",
                        help="数据目录路径")
    parser.add_argument("--user-id", default=None, help="用户 ID 筛选")
    parser.add_argument("--dashboard", action="store_true", help="生成看板")

    args = parser.parse_args()

    print("\n" + "="*70)
    print("健康数据提取与看板生成工具")
    print("="*70)

    if args.dashboard:
        # 使用一站式函数生成看板
        result, dashboard = extract_and_generate_dashboard(
            args.data_dir,
            args.user_id,
            print_output=True
        )
    else:
        # 仅提取数据
        result = extract_health_data(args.data_dir, args.user_id)

        print("\n" + "="*60)
        print("测试完成")
        print("="*60)

        # 打印回退信息
        print("\n回退信息:")
        print(generate_fallback_message(result))

        # 打印最终数据
        print("\n最终提取的数据:")
        print("生理数据:", result.physio_data)
        print("心理数据:", result.psych_data)
